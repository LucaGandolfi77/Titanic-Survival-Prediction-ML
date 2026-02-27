"""Export product data to a rich Excel workbook using *openpyxl*.

Features:

* Auto-width columns
* Bold header row with coloured background
* Alternating row colours
* Embedded product thumbnails (80 × 80 px via Pillow)
* Frozen header row + auto-filter
* Separate **Summary** sheet
"""

from __future__ import annotations

import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.logger import get_logger

log = get_logger(__name__)

# Try Pillow – only needed for thumbnail embedding
try:
    from PIL import Image as PILImage

    _PILLOW_AVAILABLE = True
except ImportError:  # pragma: no cover
    _PILLOW_AVAILABLE = False
    log.warning("Pillow not installed – images will NOT be embedded in Excel")


# ── Colour palette ──────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_ALT_FILL_A = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
_ALT_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

_THUMB_SIZE = (80, 80)
_COLUMNS = ["#", "Name", "Price", "Description", "URL", "Image_URL", "Local_Image", "Category"]


class ExcelExporter:
    """Export scraped product data into a styled ``.xlsx`` workbook.

    Args:
        output_path: Destination file path for the Excel file.
    """

    def __init__(self, output_path: str | Path) -> None:
        self._output_path = Path(output_path).resolve()
        self._output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Public API ──────────────────────────────────────────────

    def export(
        self,
        products: list[dict[str, Any]],
        source_url: str = "",
        pages_visited: int = 0,
    ) -> str:
        """Create the Excel file and return its path.

        Args:
            products: List of product dicts (keys: ``name``, ``price``,
                ``description``, ``product_url``, ``image_url``,
                ``local_image_path``).
            source_url: The base URL that was scraped.
            pages_visited: Total number of pages fetched.

        Returns:
            Absolute path to the saved ``.xlsx`` file.
        """
        wb = Workbook()

        # ── Products sheet ──────────────────────────────────────
        ws = wb.active
        if ws is None:  # pragma: no cover
            ws = wb.create_sheet()
        ws.title = "Products"
        self._write_products_sheet(ws, products)

        # ── Summary sheet ───────────────────────────────────────
        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary, products, source_url, pages_visited)

        wb.save(str(self._output_path))
        log.info("Excel exported → %s (%d products)", self._output_path, len(products))
        return str(self._output_path)

    # ── Products sheet ──────────────────────────────────────────

    def _write_products_sheet(
        self, ws: Any, products: list[dict[str, Any]]
    ) -> None:
        """Populate the *Products* worksheet."""

        # Header row
        for col_idx, header in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER

        # Data rows
        has_image_col = True
        for row_idx, prod in enumerate(products, start=2):
            fill = _ALT_FILL_A if row_idx % 2 == 0 else _ALT_FILL_B

            values = [
                row_idx - 1,
                prod.get("name", "N/A"),
                prod.get("price", "N/A"),
                prod.get("description", "N/A"),
                prod.get("product_url", "N/A"),
                prod.get("image_url", "N/A"),
                prod.get("local_image_path", ""),
                prod.get("category", ""),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 4))

            # Embed thumbnail
            local_path = prod.get("local_image_path", "")
            if local_path and _PILLOW_AVAILABLE:
                self._embed_thumbnail(ws, row_idx, col_idx=7, image_path=local_path)

        # Freeze header + auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Auto-width
        self._auto_width(ws)

    # ── Summary sheet ───────────────────────────────────────────

    def _write_summary_sheet(
        self,
        ws: Any,
        products: list[dict[str, Any]],
        source_url: str,
        pages_visited: int,
    ) -> None:
        """Populate the *Summary* worksheet."""
        images_ok = sum(1 for p in products if p.get("local_image_path"))
        total = len(products)
        success_rate = f"{images_ok / total * 100:.1f}%" if total else "0%"

        rows = [
            ("Metric", "Value"),
            ("Total products scraped", total),
            ("Total pages visited", pages_visited),
            ("Scraping timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Source URL", source_url),
            ("Products with images", images_ok),
            ("Image success rate", success_rate),
        ]
        for r_idx, (label, value) in enumerate(rows, start=1):
            lc = ws.cell(row=r_idx, column=1, value=label)
            vc = ws.cell(row=r_idx, column=2, value=value)
            if r_idx == 1:
                lc.fill = _HEADER_FILL
                lc.font = _HEADER_FONT
                vc.fill = _HEADER_FILL
                vc.font = _HEADER_FONT
            lc.border = _THIN_BORDER
            vc.border = _THIN_BORDER

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

    # ── Image embedding ─────────────────────────────────────────

    @staticmethod
    def _embed_thumbnail(ws: Any, row: int, col_idx: int, image_path: str) -> None:
        """Resize an image to 80 × 80 and embed it in the worksheet.

        Args:
            ws: Target worksheet.
            row: Row number (1-based).
            col_idx: Column number for the image.
            image_path: Local path to the image file.
        """
        try:
            p = Path(image_path)
            if not p.exists():
                return

            img = PILImage.open(p)
            img.thumbnail(_THUMB_SIZE)

            buf = BytesIO()
            fmt = img.format or "PNG"
            if fmt.upper() == "WEBP":
                fmt = "PNG"
                img = img.convert("RGBA")
            img.save(buf, format=fmt)
            buf.seek(0)

            xl_img = XlImage(buf)
            xl_img.width, xl_img.height = _THUMB_SIZE
            anchor = f"{get_column_letter(col_idx)}{row}"
            ws.add_image(xl_img, anchor)

            # Make row tall enough for the thumbnail
            ws.row_dimensions[row].height = 65

        except Exception as exc:  # noqa: BLE001
            log.warning("Could not embed image %s: %s", image_path, exc)

    # ── Auto-width ──────────────────────────────────────────────

    @staticmethod
    def _auto_width(ws: Any, min_width: int = 10, max_width: int = 60) -> None:
        """Set column widths to fit content.

        Args:
            ws: Target worksheet.
            min_width: Minimum column width in characters.
            max_width: Maximum column width in characters.
        """
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            lengths = []
            for cell in col_cells:
                if cell.value is not None:
                    lengths.append(len(str(cell.value)))
            best = max(lengths) + 3 if lengths else min_width
            ws.column_dimensions[col_letter].width = min(max(best, min_width), max_width)
