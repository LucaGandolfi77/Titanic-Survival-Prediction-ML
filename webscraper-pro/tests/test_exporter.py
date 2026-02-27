"""Unit tests for the Excel & CSV exporters.

Creates temporary product lists and verifies that files are written
correctly and contain the expected data.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

# ── Ensure project root is importable ───────────────────────────
_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from exporters.csv_exporter import CsvExporter
from exporters.excel_exporter import ExcelExporter


# ── Fixture data ────────────────────────────────────────────────

def _sample_products(n: int = 5) -> list[dict[str, Any]]:
    """Generate *n* fake product dicts."""
    return [
        {
            "name": f"Product {i}",
            "price": f"${i * 10 + 0.99:.2f}",
            "description": f"Description for product {i}.",
            "product_url": f"https://shop.example.com/product-{i}",
            "image_url": f"https://img.example.com/p{i}.jpg",
            "local_image_path": "",
            "category": "Clothing",
        }
        for i in range(1, n + 1)
    ]


# ═══════════════════════════════════════════════════════════════
# Excel exporter tests
# ═══════════════════════════════════════════════════════════════

class TestExcelExporter:
    """Tests for ``ExcelExporter``."""

    def test_creates_file(self, tmp_path: Path) -> None:
        out = tmp_path / "test.xlsx"
        ex = ExcelExporter(out)
        result = ex.export(_sample_products(3), source_url="https://example.com", pages_visited=1)
        assert Path(result).exists()

    def test_products_sheet_has_header(self, tmp_path: Path) -> None:
        out = tmp_path / "test.xlsx"
        ExcelExporter(out).export(_sample_products(2))
        wb = load_workbook(str(out))
        ws = wb["Products"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert headers[0] == "#"
        assert headers[1] == "Name"
        assert headers[2] == "Price"

    def test_products_count_matches(self, tmp_path: Path) -> None:
        products = _sample_products(4)
        out = tmp_path / "test.xlsx"
        ExcelExporter(out).export(products)
        wb = load_workbook(str(out))
        ws = wb["Products"]
        # Data rows = max_row - 1 (header)
        assert ws.max_row - 1 == 4

    def test_summary_sheet_exists(self, tmp_path: Path) -> None:
        out = tmp_path / "test.xlsx"
        ExcelExporter(out).export(_sample_products(1), source_url="https://x.com", pages_visited=2)
        wb = load_workbook(str(out))
        assert "Summary" in wb.sheetnames

    def test_summary_contains_source_url(self, tmp_path: Path) -> None:
        out = tmp_path / "test.xlsx"
        ExcelExporter(out).export(_sample_products(1), source_url="https://x.com", pages_visited=2)
        wb = load_workbook(str(out))
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 8)]
        assert "https://x.com" in values

    def test_empty_products(self, tmp_path: Path) -> None:
        out = tmp_path / "empty.xlsx"
        result = ExcelExporter(out).export([])
        assert Path(result).exists()
        wb = load_workbook(str(out))
        ws = wb["Products"]
        assert ws.max_row == 1  # header only


# ═══════════════════════════════════════════════════════════════
# CSV exporter tests
# ═══════════════════════════════════════════════════════════════

class TestCsvExporter:
    """Tests for ``CsvExporter``."""

    def test_creates_file(self, tmp_path: Path) -> None:
        out = tmp_path / "test.csv"
        result = CsvExporter(out).export(_sample_products(3))
        assert Path(result).exists()

    def test_row_count(self, tmp_path: Path) -> None:
        out = tmp_path / "test.csv"
        CsvExporter(out).export(_sample_products(5))
        with out.open("r", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
        # header + 5 data rows
        assert len(rows) == 6

    def test_header_columns(self, tmp_path: Path) -> None:
        out = tmp_path / "test.csv"
        CsvExporter(out).export(_sample_products(1))
        with out.open("r", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            fields = reader.fieldnames or []
        assert "name" in fields
        assert "price" in fields
        assert "product_url" in fields

    def test_values_correct(self, tmp_path: Path) -> None:
        products = _sample_products(2)
        out = tmp_path / "test.csv"
        CsvExporter(out).export(products)
        with out.open("r", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert rows[0]["name"] == "Product 1"
        assert rows[1]["name"] == "Product 2"

    def test_empty_products(self, tmp_path: Path) -> None:
        out = tmp_path / "empty.csv"
        result = CsvExporter(out).export([])
        assert Path(result).exists()
        with out.open("r", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
        assert len(rows) == 1  # header only
